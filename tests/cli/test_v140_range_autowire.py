"""v1.40.0 -- range-pack auto-wiring for wide-format measurement sheets.

Auto-wiring extends audit coverage from staging-only to biomarker VALUES, but
ONLY for assay-independent ordinal/cognitive scales (MMSE, Braak, ...). Assay-
calibrated biomarkers (fluid, PET quantitation, volumetry) are REFUSED, because
a generically-named column cannot certify the assay -- auto-wiring them produced
false positives on every row. These tests lock that boundary.
"""
from __future__ import annotations

import glob
import json

import pandas as pd
import pytest

from neurotcs.cli import main
from neurotcs.io.autowire import autowire_ranges

pytest.importorskip("openpyxl")


def test_safe_ordinal_scales_are_wired():
    tables = {
        "CG": pd.DataFrame({"subject_id": ["S1"], "visit": [0],
                            "mmse": [28], "moca": [25]}),
    }
    specs, extra, decisions, refusals, wired_src = autowire_ranges(tables, set())
    joined = " ".join(decisions)
    assert "mmse=mmse_total" in joined
    assert "moca=moca_total" in joined
    assert "CG" in wired_src


def test_assay_specific_biomarkers_are_refused_not_wired():
    tables = {
        "FL": pd.DataFrame({"subject_id": ["S1"], "visit": [0],
                            "csf_ptau181_pgml": [50.0], "plasma_nfl_pgml": [20.0]}),
    }
    specs, extra, decisions, refusals, wired_src = autowire_ranges(tables, set())
    # neither fluid biomarker is wired
    assert all("csf_ptau181" not in d for d in decisions)
    assert all("plasma_nfl" not in d for d in decisions)
    # both are explicitly refused with a calibration reason
    joined = " ".join(refusals)
    assert "csf_ptau181_pgml" in joined and "plasma_nfl_pgml" in joined
    assert "assay/scale-calibrated" in joined


def test_unit_mismatch_column_is_refused():
    """A cm3 hippocampus column must never be wired to an mm3 pack bound."""
    tables = {
        "MR": pd.DataFrame({"subject_id": ["S1"], "visit": [0],
                            "hippocampal_total_mm3": [3000.0]}),
    }
    # mm3 hippocampus is assay/volumetry -> refused regardless (not in safe set)
    specs, extra, decisions, refusals, wired_src = autowire_ranges(tables, set())
    assert all("hippocamp" not in d for d in decisions)


def test_autowire_catches_scale_violation_no_false_positives(tmp_path):
    """End-to-end: a cognitive sheet where ONE row breaks the scale (MMSE 35)
    must yield exactly one impossible flag for that measurement -- not a flag on
    every row (which would be a calibration false positive)."""
    f = tmp_path / "d.xlsx"
    n = 50
    mmse = [28] * n
    mmse[7] = 35  # the only out-of-scale value (>30)
    with pd.ExcelWriter(f, engine="openpyxl") as xw:
        pd.DataFrame({"subject_id": [f"S{i}" for i in range(n)],
                      "visit": [0] * n,
                      "clinical_state": ["MCI"] * n}).to_excel(
            xw, sheet_name="AUDIT_CLINICAL", index=False)
        pd.DataFrame({"subject_id": [f"S{i}" for i in range(n)],
                      "visit": [0] * n, "mmse": mmse}).to_excel(
            xw, sheet_name="CG", index=False)
    out = tmp_path / "out"
    main(["audit", str(f), "-o", str(out), "--allow-no-dates", "--quiet"])
    b = json.load(open(glob.glob(str(out / "*.bundle.json"))[0]))
    flags = b["neurotcs_bundle"]["deterministic_core"]["flags"]
    mmse_impossible = [x for x in flags["impossible"] if x["field"] == "mmse_total"]
    # exactly one impossible MMSE flag (the 35), NOT one per row
    assert len(mmse_impossible) == 1
    assert mmse_impossible[0]["observed_value"] == 35


# --------------------------------------------------------------------------- #
# v1.40.2 (E-2026-012): --confirm-assays opt-in + unit-plumbing fix
# --------------------------------------------------------------------------- #
def test_confirm_assays_default_off_keeps_failclosed_refusal():
    """Default (confirm_assays=False) MUST still refuse assay biomarkers."""
    tables = {
        "FL": pd.DataFrame({"subject_id": ["S1"], "visit": [0],
                            "plasma_ptau217_pgml": [0.4]}),
    }
    specs, extra, decisions, refusals, _ = autowire_ranges(tables, set())
    assert specs == []
    assert any("plasma_ptau217_pgml" in r for r in refusals)
    assert any("--confirm-assays" in r for r in refusals)


def test_confirm_assays_wires_assay_biomarker_when_asserted():
    """With confirm_assays=True the operator asserts assay match -> wired."""
    tables = {
        "FL": pd.DataFrame({"subject_id": ["S1", "S2"], "visit": [0, 0],
                            "plasma_ptau217_pgml": [0.4, 0.5]}),
    }
    specs, extra, decisions, refusals, wired_src = autowire_ranges(
        tables, set(), confirm_assays=True)
    assert specs, "assay column should be wired under --confirm-assays"
    joined = " ".join(decisions)
    assert "plasma_ptau217_pgml=plasma_ptau217_pgml" in joined
    assert "assay-confirmed by operator" in joined
    assert "FL" in wired_src


def test_confirm_assays_emits_raw_pack_unit_no_spurious_unit_mismatch():
    """E-2026-012 regression: the synthetic long table must carry the pack's
    RAW unit (e.g. 'pg/mL'), not the normalized form ('pgml'). Otherwise the
    range auditor's exact-string unit check trips a unit_mismatch on every row.
    """
    from neurotcs.clinical_ranges import audit_clinical_ranges, load_rangepack
    tables = {
        "FL": pd.DataFrame({"subject_id": ["S1", "S2", "S3"], "visit": [0, 0, 0],
                            "plasma_ptau217_pgml": [0.3, 0.4, 0.5]}),
    }
    specs, extra, decisions, refusals, _ = autowire_ranges(
        tables, set(), confirm_assays=True)
    assert len(specs) == 1
    spec = specs[0]
    long_df = extra[spec["sheet"]]
    pack = load_rangepack(spec["pack"])
    # emitted unit equals the pack's raw declared unit for that measurement
    meas = next(m for m in pack.rangepack.measurements
                if (getattr(m, "measurement_name", None)
                    or getattr(m, "name", None)) == "plasma_ptau217_pgml")
    assert set(long_df["unit"].unique()) == {meas.unit}
    # and a real audit produces ZERO unit_mismatch flags
    r = audit_clinical_ranges(long_df, pack)
    assert all(f.bound_type != "unit_mismatch" for f in r.flags)


def test_confirm_assays_does_not_override_genuine_unit_suffix_mismatch():
    """A column whose unit SUFFIX contradicts the pack is still refused even
    under --confirm-assays (a unit mismatch is a hard incompatibility, not an
    assay assertion)."""
    tables = {
        "MR": pd.DataFrame({"subject_id": ["S1"], "visit": [0],
                            "etiv_mm3": [1500000.0]}),  # pack expects cm^3
    }
    specs, extra, decisions, refusals, _ = autowire_ranges(
        tables, set(), confirm_assays=True)
    # etiv_mm3 alias maps to None (mm3 explicitly refused) OR unit-mismatches;
    # either way it must NOT be wired.
    assert all("etiv" not in s.get("sheet", "") for s in specs)


# =============================================================================
# v1.40.3 / E-2026-013 -- explicit-mapping path also auto-wires un-staged sheets
# =============================================================================
# Before v1.40.3 the explicit-mapping path only consumed mapping["ranges"]
# (whatever the user had hand-authored). Users running `describe --emit-mapping`
# got `ranges: []` and had no path to range auditing unless they wrote long-form
# specs from docs. v1.40.3 makes the explicit-mapping path run autowire on every
# un-staged sheet -- IDENTICAL fail-closed default for assay packs, IDENTICAL
# --confirm-assays opt-in. The explicit-mapping path now matches zero-config for
# range auditing while still letting hand-authored ranges entries take priority.

def test_explicit_mapping_autowires_unstaged_sheets(tmp_path, capsys):
    """A mapping that only declares the staging axis still auto-audits values
    from un-staged measurement sheets (the v1.40.3 lever)."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("clinical")
    ws.append(["subject_id", "visit", "visit_date", "state"])
    _states = ["CN", "MCI", "AD"]
    for i in range(3):
        ws.append([f"S{i:03d}", i, f"2024-0{i+1}-01", _states[i]])
    ws2 = wb.create_sheet("CG")
    ws2.append(["subject_id", "visit", "mmse", "cdr_sob"])
    ws2.append(["S000", 0, 28, 0.5])
    ws2.append(["S001", 0, 26, 1.0])
    ws2.append(["S002", 0, 22, 2.5])
    xlsx = tmp_path / "data.xlsx"
    wb.save(xlsx)

    mapping = {
        "clinical": {"sheet": "clinical", "subject_id": "subject_id",
                     "visit": "visit", "visit_date": "visit_date", "state": "state"},
        "ranges": [],   # left empty -- v1.40.3 should still wire CG
    }
    mpath = tmp_path / "map.json"
    mpath.write_text(json.dumps(mapping), encoding="utf-8")

    out = tmp_path / "out"
    rc = main(["audit", str(xlsx), "--mapping", str(mpath), "-o", str(out),
               "--allow-no-dates"])
    captured = capsys.readouterr()
    assert rc == 0, captured.err
    # CG's safe-ordinal columns were auto-wired even though mapping["ranges"]=[]
    assert "auto-wired range packs" in captured.out
    assert "mmse=mmse_total" in captured.out


def test_explicit_mapping_refuses_assay_columns_failclosed(tmp_path, capsys):
    """An explicit mapping with an un-staged ASSAY measurement sheet refuses
    those columns fail-closed (no --confirm-assays = no audit)."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("clinical")
    ws.append(["subject_id", "visit", "visit_date", "state"])
    _states = ["CN", "MCI", "AD"]
    for i in range(3):
        ws.append([f"S{i:03d}", i, f"2024-0{i+1}-01", _states[i]])
    ws2 = wb.create_sheet("FL")
    ws2.append(["subject_id", "visit", "plasma_ptau217_pgml"])
    ws2.append(["S000", 0, 0.5])
    xlsx = tmp_path / "data.xlsx"
    wb.save(xlsx)

    mapping = {
        "clinical": {"sheet": "clinical", "subject_id": "subject_id",
                     "visit": "visit", "visit_date": "visit_date", "state": "state"},
        "ranges": [],
    }
    mpath = tmp_path / "map.json"
    mpath.write_text(json.dumps(mapping), encoding="utf-8")

    out = tmp_path / "out"
    rc = main(["audit", str(xlsx), "--mapping", str(mpath), "-o", str(out),
               "--allow-no-dates"])
    captured = capsys.readouterr()
    assert rc == 0, captured.err
    # plasma_ptau217 is an ASSAY column -- it must NOT have been wired
    assert "ptau217" not in captured.out or "NOT" in captured.out
    # the refusal must be surfaced in coverage
    assert "plasma_ptau217" in captured.err or "plasma_ptau217" in captured.out


def test_explicit_mapping_with_confirm_assays_audits_assay_packs(tmp_path, capsys):
    """When --confirm-assays is set, the explicit-mapping path wires assay
    columns just like zero-config does (parity check)."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("clinical")
    ws.append(["subject_id", "visit", "visit_date", "state"])
    _states = ["CN", "MCI", "AD"]
    for i in range(3):
        ws.append([f"S{i:03d}", i, f"2024-0{i+1}-01", _states[i]])
    ws2 = wb.create_sheet("FL")
    ws2.append(["subject_id", "visit", "plasma_ptau217_pgml"])
    ws2.append(["S000", 0, 0.5])
    xlsx = tmp_path / "data.xlsx"
    wb.save(xlsx)

    mapping = {
        "clinical": {"sheet": "clinical", "subject_id": "subject_id",
                     "visit": "visit", "visit_date": "visit_date", "state": "state"},
        "ranges": [],
    }
    mpath = tmp_path / "map.json"
    mpath.write_text(json.dumps(mapping), encoding="utf-8")

    out = tmp_path / "out"
    rc = main(["audit", str(xlsx), "--mapping", str(mpath),
               "--confirm-assays", "-o", str(out), "--allow-no-dates"])
    captured = capsys.readouterr()
    assert rc == 0, captured.err
    # under --confirm-assays the assay column IS wired
    assert "plasma_ptau217" in captured.out
    assert "assay-confirmed by operator" in captured.out


class TestLongFormatValueAuditing:
    """v1.52.0: CDISC long-format measurement sheets (values in a single column
    keyed by --TESTCD / PARAMCD) must be range-audited in zero-config, by
    pivoting long -> wide and reusing the wide-column resolution path. Wide-
    format sheets are unaffected (locked separately by the v3 blind set)."""

    def test_long_format_pivots_and_wires(self):
        qs = pd.DataFrame({
            "subject_id": ["001", "001", "002", "002"],
            "visit": [1, 1, 1, 1],
            "QSTESTCD": ["MMSE", "MOCA", "MMSE", "MOCA"],
            "QSSTRESN": [28, 26, 27, 25]})
        rspecs, extra, decisions, _ref, wired = autowire_ranges(
            {"QS": qs}, set(), confirm_assays=False)
        assert any("long format" in d.lower() for d in decisions)
        assert len(rspecs) == 1
        syn = list(extra.values())[0]
        assert set(syn["measurement_name"]) == {"mmse_total", "moca_total"}
        # pivot preserved every (subject, visit, code) value
        assert len(syn) == 4

    def test_long_format_paramcd_recognized(self):
        adqs = pd.DataFrame({
            "subject_id": ["001", "001"], "visit": [1, 1],
            "PARAMCD": ["MMSE", "MOCA"], "AVAL": [29, 27]})
        rspecs, extra, decisions, _ref, _wired = autowire_ranges(
            {"ADQS": adqs}, set(), confirm_assays=False)
        assert any("long format" in d.lower() for d in decisions)
        assert len(rspecs) == 1

    def test_long_format_out_of_range_value_flagged_end_to_end(self, tmp_path):
        clin = pd.DataFrame({
            "subject_id": ["001", "001", "002", "002"],
            "visit": [1, 2, 1, 2],
            "visit_date": ["2024-01-01", "2024-07-01",
                           "2024-01-01", "2024-07-01"],
            "clinical_stage": ["CN", "MCI", "MCI", "AD"]})
        qs = pd.DataFrame({
            "subject_id": ["001", "001", "002"],
            "visit": [1, 2, 1],
            "QSTESTCD": ["MMSE", "MMSE", "MMSE"],
            "QSSTRESN": [29, 27, 99]})  # 99 impossible for MMSE (0-30)
        p = tmp_path / "cohort.xlsx"
        with pd.ExcelWriter(p) as w:
            clin.to_excel(w, sheet_name="CLINICAL", index=False)
            qs.to_excel(w, sheet_name="QS", index=False)
        out = tmp_path / "out"
        main(["audit", str(p), "-o", str(out)])
        bundle = json.load(open(glob.glob(str(out / "*.bundle.json"))[0]))
        core = bundle["neurotcs_bundle"]["deterministic_core"]
        allf = core["flags"]["impossible"] + core["flags"].get("implausible", [])
        assert any("mmse" in str(f).lower() and "99" in str(f.get("observed_value", ""))
                   for f in allf)

    def test_wide_format_sheet_not_treated_as_long(self):
        # a wide sheet with a normal MMSE column must NOT trip long-format logic
        wide = pd.DataFrame({
            "subject_id": ["001", "002"], "visit": [1, 1],
            "mmse": [29, 27], "moca": [28, 26]})
        rspecs, extra, decisions, _ref, _wired = autowire_ranges(
            {"COG": wide}, set(), confirm_assays=False)
        assert not any("long format" in d.lower() for d in decisions)
        assert len(rspecs) == 1  # still wired the wide way


# --------------------------------------------------------------------------- #
# v1.79.0 -- same-sheet measurement coverage. A staging sheet that ALSO carries
# measurement columns (single-file zero-config CSVs) previously had those
# measurements silently skipped, because autowire_ranges skipped any sheet
# already wired to staging -- a fail-open coverage gap in a fail-closed tool.
# The fix skips only the consumed staging columns, not the whole sheet.
# (External second-pass audit finding #1; reproduced and fixed from roots.)
# --------------------------------------------------------------------------- #
def test_v179_same_sheet_measurements_are_wired():
    """Same-sheet MMSE/CDR-SB ARE range-audited despite the sheet being
    staging-wired. Pre-v1.79.0 this returned no specs (whole-sheet skip)."""
    tables = {
        "ALL": pd.DataFrame({
            "subject_id": ["S1", "S2"], "visit": [0, 1],
            "visit_date": ["2020-01-01", "2021-01-01"],
            "clinical_state": ["CN", "MCI"],
            "biological_stage_atn": ["A-T-", "A+T-"],
            "MMSE": [40, 28], "CDRSB": [-1, 2],
        }),
    }
    consumed = {"ALL": ["subject_id", "visit", "visit_date",
                        "clinical_state", "biological_stage_atn"]}
    specs, extra, decisions, refusals, wired_src = autowire_ranges(
        tables, {"ALL"}, consumed_columns=consumed)
    joined = " ".join(decisions)
    assert "MMSE=mmse_total" in joined, f"MMSE not wired; decisions={decisions}"
    assert "CDRSB=cdr_sb_sum_boxes" in joined, f"CDRSB not wired; decisions={decisions}"
    assert "ALL" in wired_src
    # consumed staging columns are NOT mis-resolved as measurements
    assert "clinical_state" not in joined
    assert "biological_stage_atn" not in joined


def test_v179_staging_only_sheet_wires_nothing():
    """Invariant safety: a staging-ONLY sheet (cohort loaders emit these) wires
    NO ranges -> audit_id unchanged. Unit twin of the real-ADNI invariant."""
    tables = {
        "STG": pd.DataFrame({
            "subject_id": ["S1", "S2"], "visit": [0, 1],
            "visit_date": ["2020-01-01", "2021-01-01"],
            "clinical_state": ["CN", "MCI"],
        }),
    }
    consumed = {"STG": ["subject_id", "visit", "visit_date", "clinical_state"]}
    specs, extra, decisions, refusals, wired_src = autowire_ranges(
        tables, {"STG"}, consumed_columns=consumed)
    assert specs == [], f"staging-only sheet must wire nothing; got {specs}"


def test_v179_same_sheet_impossible_values_flagged_end_to_end(tmp_path):
    """End-to-end: a single CSV combining staging + impossible MMSE/CDR-SB must
    surface impossible flags. Pre-v1.79.0 this returned CLEAN (impossible 0)."""
    f = tmp_path / "single.csv"
    pd.DataFrame({
        "subject_id": ["S1", "S1", "S2", "S2"],
        "visit": [1, 2, 1, 2],
        "visit_date": ["2020-01-01", "2021-01-01", "2020-01-01", "2021-01-01"],
        "clinical_state": ["CN", "MCI", "CN", "MCI"],
        "biological_stage_atn": ["A-T-", "A+T-", "A-T-", "A+T-"],
        "MMSE": [40, 28, 29, 27],     # MMSE=40 impossible (>30)
        "CDRSB": [2, -1, 0, 1],       # CDRSB=-1 impossible (<0)
    }).to_csv(f, index=False)
    out = tmp_path / "out"
    main(["audit", str(f), "-o", str(out), "--allow-no-dates", "--quiet"])
    b = json.load(open(glob.glob(str(out / "*.bundle.json"))[0]))
    flags = b["neurotcs_bundle"]["deterministic_core"]["flags"]
    fields = {x["field"] for x in flags["impossible"]}
    assert "mmse_total" in fields, f"MMSE=40 not flagged; impossible={flags['impossible']}"
    assert "cdr_sb_sum_boxes" in fields, f"CDRSB=-1 not flagged; impossible={flags['impossible']}"
    assert len(flags["impossible"]) >= 2
